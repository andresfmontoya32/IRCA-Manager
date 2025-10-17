# ============================================================
# Procesador directo de archivos Excel por aeropuerto
# ============================================================

import os
import shutil
import pandas as pd
import openpyxl
import unicodedata
from docx import Document

# Diccionario robusto de asignación ciudad ↔ aeropuerto
aeropuerto_ciudad = {
    "ERNESTO CORTISSOZ":        "Barranquilla",
    "GUILLERMO LEÓN VALENCIA":  "Popayan",
    "EL EDÉN":                  "Armenia",
    "HACARITAMA":               "Aguachica",
    "GOLFO MORROSQUILLO":       "Tolu",
    "GERARDO TOVAR LÓPEZ":      "Buenaventura",
    "ANTONIO NARIÑO":           "Pasto",
    "JUAN CASIANO SOLÍS":       "Guapi",
    "SAN LUIS":                 "Ipiales",
    "EL EMBRUJO":               "Providencia",
    "GUSTAVO ROJAS PINILLA":    "San Andres",
    "LA FLORIDA":               "Tumaco",
    "PASTO":                    "Pasto",
    "GOLFO DE MORROSQUILLO":    "Tolu",
}

# Diccionario de abreviaturas por ciudad
abrev_ciudad = {
    "AGCA": "Aguachica",
    "ARM": "Armenia", 
    "BAQ": "Barranquilla",
    "BTURA": "Buenaventura",
    "GUAPI": "Guapi",
    "IPI": "Ipiales",
    "PASTO": "Pasto",
    "POP": "Popayan",
    "TOLU": "Tolu",
    "TUM": "Tumaco",
    "SAI": "San Andres",
    "PROV": "Providencia"
}

# Crear diccionario inverso para búsqueda por ciudad
ciudad_abrev = {v.upper(): k for k, v in abrev_ciudad.items()}

def normaliza(texto):
    texto = texto.upper().strip()
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join([c for c in texto if not unicodedata.combining(c)])
    return texto

def encuentra_plantilla(ciudad, ruta_plantillas):
    """Encuentra la plantilla Word correcta para una ciudad"""
    if not os.path.exists(ruta_plantillas):
        raise FileNotFoundError(f"No existe la carpeta de plantillas: {ruta_plantillas}")
    
    ciudad = normaliza(ciudad)
    # Buscar por abreviatura
    abrev = ciudad_abrev.get(ciudad)
    if abrev:
        nombre_plantilla = f"Plantilla_AP_{abrev}_2025.docx"
        ruta_completa = os.path.join(ruta_plantillas, nombre_plantilla)
        if os.path.exists(ruta_completa):
            return ruta_completa
    
    # Buscar de forma flexible
    for archivo in os.listdir(ruta_plantillas):
        if archivo.startswith("~$"):  # Ignorar temporales
            continue
        if archivo.endswith(".docx"):
            # Buscar coincidencias en nombre
            nombre_norm = normaliza(archivo)
            if ciudad in nombre_norm or any(palabra in nombre_norm 
                                         for palabra in ciudad.split()):
                return os.path.join(ruta_plantillas, archivo)
    
    return None

def encuentra_excel(carpeta):
    """Encuentra el archivo Excel principal en una carpeta"""
    archivos = [f for f in os.listdir(carpeta) 
                if f.endswith('.xlsx') and not f.startswith('~$')]
    if not archivos:
        return None
    
    # Preferir archivos que empiecen con 'Base'
    bases = [f for f in archivos if f.lower().startswith('base')]
    return os.path.join(carpeta, (bases[0] if bases else archivos[0]))

# === RUTAS BASE ===
ruta_origen = r"C:\Users\PAEROCIVIL\Desktop\Automatizacion\AUTOMATIZACION\2.Limpieza\Resultados_por_Aeropuerto"
ruta_destino_raiz = r"C:\Users\PAEROCIVIL\Desktop\Automatizacion\AUTOMATIZACION\3.correspondencia\Datos"
ruta_plantillas = r"C:\Users\PAEROCIVIL\Desktop\Automatizacion\AUTOMATIZACION\3.correspondencia\Plantillas"

print("🔍 Verificando ruta de origen...")
if not os.path.exists(ruta_origen):
    print(f"❌ ERROR: La carpeta de origen NO existe: {ruta_origen}")
    exit()
else:
    print(f"✅ Carpeta de origen encontrada: {ruta_origen}")

print("🔍 Verificando ruta de plantillas Word...")
if not os.path.exists(ruta_plantillas):
    print(f"❌ ERROR: La carpeta de plantillas NO existe: {ruta_plantillas}")
    exit()
else:
    print(f"✅ Carpeta de plantillas encontrada: {ruta_plantillas}")

archivos = os.listdir(ruta_origen)
print(f"📦 Archivos encontrados en '{ruta_origen}':")
if not archivos:
    print("⚠️ ADVERTENCIA: La carpeta está vacía.")
else:
    for archivo in archivos:
        print(f"   • {archivo}")

# === CREAR CARPETA DE DESTINO SI NO EXISTE ===
os.makedirs(ruta_destino_raiz, exist_ok=True)
print(f"✅ Carpeta destino preparada: {ruta_destino_raiz}")

# === PROCESAR ARCHIVOS ===
archivos_procesados = 0
for archivo in archivos:
    if archivo.lower().endswith(".xlsx"):
        # Extraer nombre del aeropuerto eliminando prefijos y extensión
        aeropuerto = (archivo
                     .replace("Base_Aeropuerto ", "")
                     .replace("base_aeropuerto ", "")
                     .replace(".xlsx", "")
                     .strip())
        
        # Buscar correspondencia en el diccionario
        aeropuerto_normalizado = normaliza(aeropuerto)
        ciudad = None
        
        # Intentar encontrar coincidencia en el diccionario
        for key, value in aeropuerto_ciudad.items():
            if normaliza(key) in aeropuerto_normalizado or aeropuerto_normalizado in normaliza(key):
                ciudad = value
                break
        
        if ciudad is None:
            print(f"⚠️ ADVERTENCIA: No se encontró correspondencia para el aeropuerto '{aeropuerto}'")
            ciudad = aeropuerto
        
        print(f"\n📄 Procesando: {archivo}")
        print(f"   🔄 Aeropuerto: {aeropuerto} → Ciudad: {ciudad}")

        # Crear estructura de carpetas y copiar archivo
        carpeta_ciudad = os.path.join(ruta_destino_raiz, ciudad)
        os.makedirs(carpeta_ciudad, exist_ok=True)
        print(f"   📁 Carpeta creada: {carpeta_ciudad}")

        # Renombrar archivo destino usando el nombre de la ciudad
        nombre_excel_ciudad = f"base_{ciudad}.xlsx"
        ruta_excel_origen = os.path.join(ruta_origen, archivo)
        ruta_excel_destino = os.path.join(carpeta_ciudad, nombre_excel_ciudad)

        try:
            shutil.copyfile(ruta_excel_origen, ruta_excel_destino)
            print(f"   ✅ Copiado: {archivo} → {nombre_excel_ciudad}")
        except Exception as e:
            print(f"❌ ERROR al copiar archivo Excel: {e}")
            continue

        # === PROCESAR DIRECTAMENTE EL EXCEL ===
        ruta_excel = ruta_excel_destino
        hoja_origen = "Sheet1"
        hoja_salida = "TABLA_4"

        print("🚀 Iniciando generación de TABLA_4 para el archivo:", ruta_excel)

        if not os.path.exists(ruta_excel):
            print(f"❌ ERROR: No se encontró el archivo '{ruta_excel}'.")
            continue

        # === LECTURA DE DATOS ===
        try:
            print(f"📥 Leyendo hoja '{hoja_origen}' del archivo...")
            df = pd.read_excel(ruta_excel, sheet_name=hoja_origen)
            print("✅ Hoja leída con éxito.")
        except Exception as e:
            print(f"❌ ERROR al leer la hoja '{hoja_origen}'. Detalle: {e}")
            continue

        # === NORMALIZACIÓN DE COLUMNAS ===
        df.columns = df.columns.str.strip().str.upper()
        df.rename(columns={"PUNTO DE MUESTREO": "PUNTO", "MÉTODO": "METODO"}, inplace=True)

        print(f"📊 Columnas encontradas: {df.columns.tolist()}")

        requeridas = ['PARÁMETRO', 'TÉCNICA', 'UNIDAD', 'LÍMITE', 'PUNTO', 'RESULTADO_CRUDO']
        faltan = [c for c in requeridas if c not in df.columns]
        if faltan:
            print(f"❌ ERROR: faltan columnas: {faltan}")
            continue

        df = df[requeridas]
        print("🔍 Vista previa datos originales:\n", df.head())

        # ==========================================================
        #        *** PRESERVAR EL ORDEN DE LOS PARÁMETROS ***
        # ==========================================================
        # 1️⃣ Guardar el orden original de aparición de 'PARÁMETRO'
        orden_param = (
            df.drop_duplicates(subset=['PARÁMETRO'])['PARÁMETRO']
              .tolist()
        )

        # 2️⃣ Orden de los puntos tal como aparecen (sin sorted)
        puntos_unicos = df['PUNTO'].dropna().unique().tolist()

        # 3️⃣ Pivot sin ordenar
        try:
            tabla = df.pivot_table(
                index=['PARÁMETRO', 'TÉCNICA', 'UNIDAD', 'LÍMITE'],
                columns='PUNTO',
                values='RESULTADO_CRUDO',
                aggfunc='first',
                sort=False
            ).reset_index()
            print("✅ Pivot exitoso.")
        except Exception as e:
            print(f"❌ ERROR en pivot: {e}"); continue

        # 4️⃣ Reordenar según 'orden_param'
        tabla['__orden__'] = pd.Categorical(tabla['PARÁMETRO'], categories=orden_param, ordered=True)
        tabla = (tabla.sort_values('__orden__')
                       .drop(columns='__orden__')
                       .reset_index(drop=True))

        # 5️⃣ Ajustar columnas finales y valores nulos
        columnas_finales = ['PARÁMETRO', 'TÉCNICA', 'UNIDAD', 'LÍMITE'] + list(puntos_unicos)
        tabla = tabla.reindex(columns=columnas_finales, fill_value="n/a")

        print("📄 Vista previa de TABLA_4 ordenada:\n", tabla.head())

        # === CREAR HOJA TAGS VACÍA ====================================
        print("🧩 Generando hoja 'TAGS'...")

        etiquetas = [
            'nro', 'periodo', 'mes', 'año', 'pto_1', 'pto_2', 'pto_3', 'pto_4',
            'fecha_mu', 'dia_mu', 'cod_1', 'cod_2', 'cod_3', 'cod_4',
            'irca_pto_1', 'irca_pto_2', 'irca_pto_3', 'irca_pto_4',"clasificacion_riesgo_1", 
            "clasificacion_riesgo_2","clasificacion_riesgo_3", "clasificacion_riesgo_4","param_1", 'param_2', 'param_3', 
            'param_4',"param_5", 'param_6'
        ]

        # Leer hoja TAGS si existe, para no perder datos manuales
        try:
            df_tags_existente = pd.read_excel(ruta_excel, sheet_name='TAGS')
        except Exception:
            df_tags_existente = pd.DataFrame(columns=['ETIQUETA', 'VALOR'])

        # Crear DataFrame base con todas las etiquetas necesarias
        df_tags_nuevo = pd.DataFrame({'ETIQUETA': etiquetas})

        # Unir con los valores existentes (sin duplicar etiquetas)
        df_tags = pd.merge(
            df_tags_nuevo,
            df_tags_existente,
            on='ETIQUETA',
            how='left'
        )
        df_tags['VALOR'] = df_tags['VALOR'].combine_first(df_tags['VALOR'])
        df_tags = df_tags[['ETIQUETA', 'VALOR']]

        # === ESCRIBIR PRIMERO LA HOJA TAGS EN EL EXCEL ===
        try:
            with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='w') as writer:
                df_tags.to_excel(writer, sheet_name='TAGS', index=False)
            print(f"✅ Hoja 'TAGS' creada como primera hoja.")
        except Exception as e:
            print(f"❌ ERROR al crear hoja TAGS: {e}")

        # === ESCRIBIR TABLA_4 EN EL EXCEL ===
        try:
            with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                tabla.to_excel(writer, sheet_name=hoja_salida, index=False)
            print(f"✅ Hoja '{hoja_salida}' escrita.")
        except Exception as e:
            print(f"❌ ERROR al escribir hoja TABLA_4: {e}"); continue

        # === CAMBIAR FUENTE DE TABLA_4 Y TAGS A VERDANA 6pt ===
        try:
            wb = openpyxl.load_workbook(ruta_excel)
            from openpyxl.styles import Font
            font = Font(name='Verdana', size=6)
            for hoja in [hoja_salida, 'TAGS']:
                if hoja in wb.sheetnames:
                    ws = wb[hoja]
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.font = font
                    print(f"✅ Fuente de hoja '{hoja}' cambiada a Verdana 6pt.")
                else:
                    print(f"❌ ERROR: Hoja '{hoja}' no encontrada para cambiar fuente.")
            wb.save(ruta_excel)
        except Exception as e:
            print(f"❌ ERROR al cambiar fuente en hojas TABLA_4/TAGS: {e}")

        # === CREAR HOJA TAGS VACÍA ====================================
        print("🧩 Generando hoja 'TAGS'...")

        etiquetas = [
            'nro', 'periodo', 'mes', 'año', 'pto_1', 'pto_2', 'pto_3', 'pto_4',
            'fecha_mu', 'dia_mu', 'cod_1', 'cod_2', 'cod_3', 'cod_4',
            'irca_pto_1', 'irca_pto_2', 'irca_pto_3', 'irca_pto_4',"clasificacion_riesgo_1", 
            "clasificacion_riesgo_2","clasificacion_riesgo_3", "clasificacion_riesgo_4","param_1", 'param_2', 'param_3', 
            'param_4',"param_5", 'param_6'
        ]

        # Leer hoja TAGS si existe, para no perder datos manuales
        try:
            df_tags_existente = pd.read_excel(ruta_excel, sheet_name='TAGS')
        except Exception:
            df_tags_existente = pd.DataFrame(columns=['ETIQUETA', 'VALOR'])

        # Crear DataFrame base con todas las etiquetas necesarias
        df_tags_nuevo = pd.DataFrame({'ETIQUETA': etiquetas})

        # Unir con los valores existentes (sin duplicar etiquetas)
        df_tags = pd.merge(
            df_tags_nuevo,
            df_tags_existente,
            on='ETIQUETA',
            how='left'
        )
        df_tags['VALOR'] = df_tags['VALOR'].combine_first(df_tags['VALOR'])
        df_tags = df_tags[['ETIQUETA', 'VALOR']]

        try:
            with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_tags.to_excel(writer, sheet_name='TAGS', index=False)
            print("✅ Hoja 'TAGS' creada o actualizada sin duplicar ni borrar datos existentes.")
        except Exception as e:
            print(f"❌ ERROR al crear hoja TAGS: {e}")

        # Después de crear la carpeta de ciudad
        # Buscar y copiar plantilla Word si no existe
        plantilla = encuentra_plantilla(ciudad, ruta_plantillas)
        if plantilla:
            nombre_plantilla = os.path.basename(plantilla)
            ruta_plantilla_destino = os.path.join(carpeta_ciudad, nombre_plantilla)
            if not os.path.exists(ruta_plantilla_destino):
                try:
                    shutil.copy2(plantilla, ruta_plantilla_destino)
                    print(f"   ✅ Plantilla Word copiada: {nombre_plantilla}")
                except Exception as e:
                    print(f"❌ ERROR al copiar plantilla Word: {e}")
        else:
            print(f"⚠️ ADVERTENCIA: No se encontró plantilla Word para {ciudad}")

        archivos_procesados += 1

# === RESUMEN FINAL ===
print(f"\n🏁 Proceso terminado. Ciudades procesadas: {archivos_procesados}")
if archivos_procesados == 0:
    print("⚠️ ADVERTENCIA: No se procesaron archivos Excel válidos.")